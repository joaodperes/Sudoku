from datetime import datetime

board = []

base  = 3
side  = base*base

# pattern for a baseline valid solution
def pattern(r,c): return (base*(r%base)+r//base+c)%side

# randomize rows, columns and numbers (of valid base pattern)
from random import sample
def shuffle(s): return sample(s,len(s)) 
rBase = range(base) 
rows  = [ g*base + r for g in shuffle(rBase) for r in shuffle(rBase) ] 
cols  = [ g*base + c for g in shuffle(rBase) for c in shuffle(rBase) ]
nums  = shuffle(range(1,base*base+1))

# produce board using randomized baseline pattern
board1 = [ [nums[pattern(r,c)] for c in cols] for r in rows ]

#for line in board: print(line)

squares = side*side
empties = squares * 3//4
for p in sample(range(squares),empties):
    board1[p//side][p%side] = " "

numSize = len(str(side))
for line in board1:
    #print(line)
    board.append(line)

def print_board(bo, file):

    for i in range(len(bo)):
        if i % 3 == 0 and i != 0:
            print("-----------------------", file=open(file_name, 'a'))

        for j in range(len(bo[0])):
            if j % 3 == 0 and j != 0:
                    print(" | ", end="", file=open(file_name, 'a'))

            if j == 8:
                print(bo[i][j], file=open(file_name, 'a'))
            else:
                print(str(bo[i][j]) + " ", end="", file=open(file_name, 'a'))

def find_empty(bo):

    for i in range(len(bo)):
        for j  in range(len(bo[0])):
            if bo[i][j] == 0:
                return (i, j) #row, col

    return None

def valid(bo, num, pos):
    # check row
    for i in range(len(bo[0])):
        if bo[pos[0]][i] == num and pos[1] != 1:
            return False

    # check column
    for i in range(len(bo)):
        if bo[i][pos[1]] == num and pos[0] != 1:
            return False

    # check box
    box_x = pos[1] // 3
    box_y = pos[0] // 3

    for i in range(box_y * 3, box_y * 3 + 3):
        for j in range(box_x * 3, box_x * 3 + 3):
            if bo[i][j] == num and (i,j) != pos:
                return False

    return True

def solve(bo):

    find = find_empty(bo)
    if not find:
        return True
    else:
        row, col = find

    for i in range(1,10):
        if valid(bo, i, (row, col)):
            bo[row][col] = i

            if solve(bo):
                return True

            bo[row][col] = 0

    return False

t = datetime.now()
file_name = t.strftime('%d-%m-%y.txt')
file_name = 'C:\\Python\\Sudoku\\' + file_name
with open(file_name, 'w'):
    print_board(board, file_name)
file1 = file_name

solve(board)
file_name = t.strftime('%d-%m-%y.txt')
file_name = 'C:\\Python\\Sudoku\\' + 'Solution_' + file_name
with open(file_name, 'w') as sd:
    print_board(board, file_name)

file2 = file_name

import yagmail
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('C:\\Python\\Sudoku\\Contacts_Sudoku.xlsx')
ws = wb.active

for row in range(2,len(ws['B'])+1):
    for col in range(2,3):
        char = get_column_letter(col)
        #print(ws[char + str(row)].value)
        yag = yagmail.SMTP('Your-Email-Address-Here', oauth2_file='C:\\Python\\GoogleAPI\\credentials.json')
        subject = [
            "Daily Sudoku - " + t.strftime('%d-%m-%y')
        ]
        contents = [
            "Hi,",
            "Please find the daily sudoku and its solution attached to this email.",
            "",
            "Good luck!"
        ]
        yag.send(ws[char + str(row)].value, subject, contents, attachments=[file1, file2])

wb.save('C:\\Python\\Sudoku\\Contacts_Sudoku.xlsx')